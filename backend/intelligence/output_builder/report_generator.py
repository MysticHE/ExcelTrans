"""
Generates Excel comparison reports using openpyxl.
Supports multiple rule output columns, column filtering, and custom column order.
"""
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Any, Dict, List, Optional
from datetime import datetime


def _hex_to_fill(hex_color: Optional[str]) -> Optional[PatternFill]:
    """Convert #RRGGBB to openpyxl PatternFill. Always create new Fill objects."""
    if not hex_color:
        return None
    color = hex_color.lstrip('#')
    if len(color) == 6:
        return PatternFill(start_color=color, end_color=color, fill_type='solid')
    return None


def _make_header_fill() -> PatternFill:
    return PatternFill(start_color='366092', end_color='366092', fill_type='solid')


def _make_header_font() -> Font:
    return Font(bold=True, color='FFFFFF', name='Calibri', size=11)


def _make_summary_title_fill() -> PatternFill:
    return PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')


def build_comparison_report(
    result: Dict[str, Any],
    template_name: str,
    compare_fields: List[str],
    key_fields: List[str],
    display_fields: List[str],
    add_remarks: bool = True,
    include_summary: bool = True,
    highlight_changed_cells: bool = True,
    output_sheet_name: str = 'Comparison',
    included_columns: Optional[List[str]] = None,
    column_order: Optional[List[str]] = None,
) -> bytes:
    """
    Build an Excel workbook from the rule executor result dict.
    Supports multiple rule output columns, included_columns filter, and column_order.

    Returns raw bytes of the .xlsx file.
    """
    wb = openpyxl.Workbook()

    # ── Main comparison sheet ──────────────────────────────────────────────────
    ws = wb.active
    safe_name = (output_sheet_name or 'Comparison')[:31]
    ws.title = safe_name

    # Collect all data fields (key + display + compare)
    all_data_fields = list(dict.fromkeys(key_fields + display_fields + compare_fields))

    # Collect extra rule output column names from row data
    extra_rule_cols: List[str] = []
    seen_extra: set = set()
    for row in result.get('rows', []):
        for col_name in row.get('output_columns', {}):
            if col_name == 'Remarks':
                continue
            if col_name not in seen_extra:
                extra_rule_cols.append(col_name)
                seen_extra.add(col_name)

    # Build full header list: data fields + Remarks + extra rule columns
    all_headers: List[str] = list(all_data_fields)
    if add_remarks:
        all_headers.append('Remarks')
    for ec in extra_rule_cols:
        all_headers.append(ec)

    # Apply included_columns filter
    if included_columns is not None:
        included_set = set(included_columns)
        all_headers = [h for h in all_headers if h in included_set]

    # Apply column_order reordering
    if column_order:
        order_map = {col: i for i, col in enumerate(column_order)}
        def _sort_key(h: str) -> int:
            return order_map.get(h, len(column_order))
        all_headers = sorted(all_headers, key=_sort_key)

    # Track max column widths during write (avoids a second O(rows×cols) pass)
    col_widths = [max(10, len(h)) for h in all_headers]

    # Write headers
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _make_header_fill()
        cell.font = _make_header_font()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write data rows — track column widths inline
    for row_idx, row_data in enumerate(result.get('rows', []), start=2):
        row_color = row_data.get('color')
        changed_fields = set(row_data.get('changed_fields', []))
        row_fill = _hex_to_fill(row_color)
        output_cols_data = row_data.get('output_columns', {})

        for col_idx, field in enumerate(all_headers, start=1):
            # Determine cell value based on field type
            if field == 'Remarks':
                value = output_cols_data.get('Remarks', {}).get('label', row_data.get('remarks', ''))
            elif field in seen_extra:
                value = output_cols_data.get(field, {}).get('label', '')
            else:
                value = row_data.get(field)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center', wrap_text=False)

            # Track column width
            if value is not None:
                col_widths[col_idx - 1] = min(40, max(col_widths[col_idx - 1], len(str(value))))

            # Row-level coloring
            if row_fill:
                cell.fill = PatternFill(
                    start_color=row_fill.start_color.rgb,
                    end_color=row_fill.end_color.rgb,
                    fill_type='solid'
                )

            # Override with cell-level change highlight if applicable
            if highlight_changed_cells and field in changed_fields and row_color is None:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    # Apply collected column widths
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

    ws.freeze_panes = 'A2'

    # ── Summary sheet ──────────────────────────────────────────────────────────
    if include_summary:
        summary_data = result.get('summary', {})
        ws_sum = wb.create_sheet(title='Summary')

        ws_sum.cell(row=1, column=1, value=f'{template_name} — Comparison Summary').font = Font(bold=True, size=14)
        ws_sum.cell(row=2, column=1, value=f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}')

        summary_rows = [
            ('Total rows processed', summary_data.get('total', 0)),
            ('Additions (new in File B)', summary_data.get('additions', 0)),
            ('Deletions (removed from File A)', summary_data.get('deletions', 0)),
            ('Changes (modified fields)', summary_data.get('changes', 0)),
            ('Unchanged', summary_data.get('unchanged', 0)),
        ]

        for i, (label, val) in enumerate(summary_rows, start=4):
            label_cell = ws_sum.cell(row=i, column=1, value=label)
            ws_sum.cell(row=i, column=2, value=val)
            label_cell.fill = _make_summary_title_fill()
            label_cell.font = Font(bold=True)

        ws_sum.column_dimensions['A'].width = 40
        ws_sum.column_dimensions['B'].width = 15

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
