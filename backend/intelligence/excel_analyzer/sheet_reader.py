"""
Dual-pass openpyxl reader: reads formulas AND computed values from Excel sheets.
"""
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from typing import Any, Dict, List, Optional, Tuple


def read_sheet_dual_pass(filepath: str) -> Dict[str, Any]:
    """
    Perform two openpyxl loads:
    1. data_only=False → capture raw formulas
    2. data_only=True  → capture computed values (cached by Excel)

    Returns a dict keyed by sheet name with cell-level data merged.
    """
    formula_wb = openpyxl.load_workbook(filepath, data_only=False)
    value_wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    sheets = {}
    for sheet_name in formula_wb.sheetnames:
        f_ws = formula_wb[sheet_name]
        v_ws = value_wb[sheet_name] if sheet_name in value_wb.sheetnames else None

        rows = []
        v_row_iter = v_ws.iter_rows() if v_ws is not None else None
        for row_idx, f_row in enumerate(f_ws.iter_rows(), start=1):
            v_row = next(v_row_iter) if v_row_iter is not None else None
            v_vals = {vc.column: vc.value for vc in v_row} if v_row else {}

            row_data = []
            for cell in f_row:
                # MergedCell placeholders cover non-anchor positions of merged ranges.
                # They have no value, formula, or column_letter — treat as empty.
                if isinstance(cell, MergedCell):
                    row_data.append({
                        'row': row_idx,
                        'col': cell.column,
                        'col_letter': get_column_letter(cell.column),
                        'formula': None,
                        'value': None,
                        'is_formula': False,
                        'data_type': None,
                    })
                    continue

                formula_val = cell.value
                is_formula = isinstance(formula_val, str) and formula_val.startswith('=')

                # Get computed value from second pass via O(1) column lookup
                computed = v_vals.get(cell.column)

                row_data.append({
                    'row': row_idx,
                    'col': cell.column,
                    'col_letter': cell.column_letter,
                    'formula': formula_val if is_formula else None,
                    'value': computed if is_formula else formula_val,
                    'is_formula': is_formula,
                    'data_type': cell.data_type,
                })
            rows.append(row_data)

        sheets[sheet_name] = {
            'rows': rows,
            'max_row': f_ws.max_row,
            'max_col': f_ws.max_column,
            'merged_cells': [str(m) for m in f_ws.merged_cells.ranges],
        }

    formula_wb.close()
    value_wb.close()
    return sheets


def extract_values_grid(sheet_data: Dict[str, Any]) -> List[List[Any]]:
    """Extract just the computed values as a 2D grid."""
    return [[cell['value'] for cell in row] for row in sheet_data['rows']]


def extract_formula_grid(sheet_data: Dict[str, Any]) -> List[List[Optional[str]]]:
    """Extract formula strings as a 2D grid (None for non-formula cells)."""
    return [[cell['formula'] for cell in row] for row in sheet_data['rows']]


def get_formula_cells(sheet_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Return a flat list of all formula cells with their location and formula string."""
    formula_cells = []
    for row in sheet_data['rows']:
        for cell in row:
            if cell['is_formula']:
                formula_cells.append({
                    'row': cell['row'],
                    'col': cell['col'],
                    'col_letter': cell['col_letter'],
                    'formula': cell['formula'],
                    'value': cell['value'],
                })
    return formula_cells
